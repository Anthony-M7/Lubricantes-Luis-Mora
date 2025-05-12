from django.http import JsonResponse
from ..models import Articulo, MovimientoInventario, HistorialStock, Categoria
from barcode import get_barcode_class
from barcode.writer import ImageWriter
from io import BytesIO
import base64
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from ..forms import CompraInventarioForm, ArticuloForm
from django.db import transaction
import os
from django.db.models import ProtectedError

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from weasyprint import HTML
from django.conf import settings
import os
from django.db import models

from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# =================================================================================

def producto_detalle_api(request, pk):
    try:
        producto = Articulo.objects.get(pk=pk)

        # Obtener última compra (ENTRADA)
        ultima_compra = MovimientoInventario.objects.filter(
            articulo=producto,
            tipo='ENTRADA'
        ).order_by('-fecha').first()

        # Obtener última venta (SALIDA)
        ultima_venta = MovimientoInventario.objects.filter(
            articulo=producto,
            tipo='SALIDA'
        ).order_by('-fecha').first()

        # Generar código de barras
        barcode_buffer = BytesIO()
        if producto.codigo_barras:
            BarcodeClass = get_barcode_class('code128')
            codigo = BarcodeClass(str(producto.codigo_barras), writer=ImageWriter())
            codigo.write(
                barcode_buffer,
                options={'write_text': False, 'module_height': 10}
            )
            barcode_base64 = base64.b64encode(barcode_buffer.getvalue()).decode('utf-8')
        else:
            barcode_base64 = None
        
        # Preparar datos de última compra
        ultima_compra_data = None
        if ultima_compra:
            ultima_compra_data = {
                'fecha': ultima_compra.fecha.strftime("%d/%m/%Y %H:%M"),
                'cantidad': int(ultima_compra.cantidad),
                'precio_unitario': int(ultima_compra.costo_unitario) if ultima_compra.costo_unitario else int(producto.costo_promedio),
                'referencia': ultima_compra.referencia,
                'total': int(ultima_compra.cantidad * ultima_compra.costo_unitario),
                'usuario': ultima_compra.usuario.get_full_name()
            }

        # Preparar datos de última venta
        ultima_venta_data = None
        if ultima_venta:
            ultima_venta_data = {
                'fecha': ultima_venta.fecha.strftime("%d/%m/%Y %H:%M"),
                'cantidad': int(ultima_venta.cantidad),
                'precio_unitario': int(producto.precio_venta),  # Usamos el precio actual del artículo
                'referencia': ultima_venta.referencia,
                'total': int(ultima_venta.cantidad * ultima_venta.costo_unitario),
                'usuario': ultima_venta.usuario.get_full_name()
            }

        data = {
            'id': producto.id,
            'codigo': producto.codigo,
            'codigo_barras': producto.codigo_barras,
            'nombre': producto.nombre,
            'descripcion': producto.descripcion,
            'precio_venta': str(producto.precio_venta),
            'costo_promedio': str(producto.costo_promedio),
            'stock_actual': str(producto.stock_actual),
            'stock_minimo': str(producto.stock_minimo),
            'stock_maximo': str(producto.stock_maximo) if producto.stock_maximo else None,
            'unidad_medida': producto.get_unidad_medida_display(),
            'categoria': producto.categoria.nombre if producto.categoria else '',
            'marca': producto.marca if producto.marca else 'No especificada',
            'modelo': producto.modelo if producto.modelo else 'No especificado',
            'proveedor': producto.proveedor_principal.nombre if producto.proveedor_principal else 'No asignado',
            'impuesto': f"{float(producto.tasa_impuesto)*100}%",
            'imagen': producto.imagen.url if producto.imagen else None,
            'palabras_clave': producto.palabras_clave,
            'barcode': barcode_base64,
            'necesita_reabastecimiento': producto.necesita_reabastecimiento,
            'lead_time': producto.lead_time,
            'activo': producto.activo,
            'valor_total': str(producto.valor_total),
            'margen_ganancia': f"{producto.margen_ganancia:.2f}%",
            'creado_por': producto.creado_por.get_full_name() if producto.creado_por else 'Sistema',
            'ultima_actualizacion': producto.fecha_actualizacion.strftime("%d/%m/%Y %H:%M"),
            'fecha_creacion': producto.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
            'ultima_compra': ultima_compra_data,
            'ultima_venta': ultima_venta_data
        }
        return JsonResponse(data)
    except Articulo.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)


def api_articulos(request):
    # Obtener solo artículos activos
    articulos = Articulo.objects.filter(activo=True).values(
        'id',
        'nombre',
        'codigo',
        'unidad_medida',
        'stock_actual',
        'costo_promedio',
        "precio_venta",
    )
    
    # Convertir QuerySet a lista
    articulos_list = list(articulos)
    
    # Retornar como JSON
    return JsonResponse(articulos_list, safe=False)


def eliminar_compra(request, pk):
    # Obtener el movimiento de compra
    movimiento = get_object_or_404(MovimientoInventario, pk=pk, tipo='ENTRADA')
    
    if request.method == 'POST':
        try:
            articulo = movimiento.articulo
            
            # 3. Eliminar el movimiento
            movimiento.delete()
            
            messages.success(request, 'Compra eliminada correctamente y stock revertido')
            
        except Exception as e:
            messages.error(request, f'Error al eliminar la compra: {str(e)}')
            return redirect('compras')
    
    return redirect('compras')


def detalle_compra(request, pk):
    # Obtener datos principales
    compra = get_object_or_404(
        MovimientoInventario.objects.select_related('articulo', 'proveedor', 'usuario'),
        pk=pk,
        tipo='ENTRADA'
    )
    historico = get_object_or_404(HistorialStock, movimiento=compra)
    
    # Calcular valores necesarios
    total = compra.cantidad * compra.costo_unitario
    
    # Variación de cantidad (% aumento stock)
    aumento_stock_pct = (compra.cantidad / historico.stock_antes * 100) if historico.stock_antes > 0 else 0
    aumento_stock_text = f"+{compra.cantidad:.0f} ({aumento_stock_pct:.1f}%)"
    
    # Variación de precio (vs compra anterior)
    if historico.costo_unitario_anterior:
        variacion_precio = (
            (compra.costo_unitario - historico.costo_unitario_anterior) / 
            historico.costo_unitario_anterior * 100
        )
    else:
        variacion_precio = 0
    
    # Variación de costo promedio
    if historico.costo_promedio_antes > 0:
        variacion_promedio = (
            (historico.costo_promedio_despues - historico.costo_promedio_antes) / 
            historico.costo_promedio_antes * 100
        )
        variacion_promedio_text = f"{'↑' if variacion_promedio > 0 else '↓'} {abs(variacion_promedio):.1f}%"
    else:
        variacion_promedio = 0
        variacion_promedio_text = "0.0%"
    
    context = {
        'compra': compra,
        'historico': historico,
        'titulo': f'Detalle de Compra #{compra.id}',
        'total': total,
        'aumento_stock_text': aumento_stock_text,
        'variacion_precio': variacion_precio,
        'variacion_promedio_text': variacion_promedio_text,
    }
    
    return render(request, 'detalles/detalles_compra.html', context)


def editar_compra(request, pk):
    compra = get_object_or_404(MovimientoInventario, pk=pk, tipo='ENTRADA')
    
    if request.method == 'POST':
        form = CompraInventarioForm(request.POST, instance=compra, user=request.user)
        if form.is_valid():
            # El formulario se encarga de toda la lógica de actualización
            form.save()
            messages.success(request, 'Compra actualizada correctamente')
            return redirect('detalle_compra', pk=compra.pk)
    else:
        form = CompraInventarioForm(instance=compra, user=request.user)
    
    return render(request, 'Forms/editar_compra.html', {
        'form': form,
        'compra': compra,
        "titulo": f"Editar Compra #{compra.id}",
        "codigo": compra.articulo.codigo,
        "articulo": compra.articulo.nombre,
        "stock_actual": int(compra.articulo.stock_actual),
        "costo_promedio": compra.articulo.costo_promedio,
    })


def editar_producto(request, producto_id):
    producto = get_object_or_404(Articulo, id=producto_id)
    categorias = Categoria.objects.all()
    
    # Obtener últimos movimientos
    movimientos = MovimientoInventario.objects.filter(articulo=producto).order_by('-fecha')[:5]
    
    if request.method == 'POST':
        form = ArticuloForm(request.POST, request.FILES, instance=producto, editing=True, user=request.user)
        if form.is_valid():
            producto = form.save(commit=False)
            
            # Restauramos los valores de los campos readonly que Django no guarda
            producto.codigo = Articulo.objects.get(id=producto_id).codigo
            producto.codigo_barras = Articulo.objects.get(id=producto_id).codigo_barras
            producto.creado_por = Articulo.objects.get(id=producto_id).creado_por
            producto.fecha_creacion = Articulo.objects.get(id=producto_id).fecha_creacion
            
            # Manejar eliminación de imagen si se seleccionó
            if 'eliminar_imagen' in request.POST and request.POST['eliminar_imagen'] == 'on':
                producto.imagen.delete(save=True)
            
            producto.save()
            messages.success(request, f'El producto "{producto.nombre}" ha sido actualizado correctamente.')
            return redirect('inventario')
    else:
        form = ArticuloForm(instance=producto, editing=True, user=request.user)
    
    context = {
        'producto': producto,
        'form': form,
        'categorias': categorias,
        'movimientos': movimientos,
    }
    return render(request, 'Forms/editar_articulos.html', context)


def eliminar_producto(request, producto_id):
    producto = get_object_or_404(Articulo, id=producto_id)
    nombre_producto = producto.nombre

    if request.method == 'POST':
        # Eliminar imagen si existe
        if producto.imagen and os.path.isfile(producto.imagen.path):
            os.remove(producto.imagen.path)

        try:
            producto.delete()
            messages.success(request, f'El producto "{nombre_producto}" ha sido eliminado correctamente.')
        except ProtectedError:
            # Si está en uso, desactivar en lugar de eliminar
            producto.activo = False
            producto.save()
            messages.warning(
                request,
                f'El producto "{nombre_producto}" no se eliminó porque está en uso, pero fue desactivado.'
            )

        return redirect('inventario')

    return redirect('editar_producto', producto_id=producto_id)

# =================================================================================

def generar_catalogo_pdf(request):
    # Obtener artículos organizados por categoría
    categorias = Categoria.objects.prefetch_related(
        models.Prefetch(
            'articulo_set',
            queryset=Articulo.objects.filter(activo=True).select_related('proveedor_principal'),
            to_attr='articulos'
        )
    ).order_by('nombre')

    # Contexto para el template
    context = {
        'categorias': categorias,
        'empresa': "Lubricantes Luis Mora",
        'fecha': timezone.now().strftime("%d/%m/%Y"),
        'base_url': request.build_absolute_uri('/')[:-1]  # Para imágenes
    }

    # Renderizar template HTML
    html_string = render_to_string('catalogo_pdf.html', context)
    
    # Crear respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename="catalogo_lubricantes.pdf"'
    
    # Generar PDF
    HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf(response)
    
    return response

# ==================================================================================

class GenerarTalonarioPagoExcel(View):
    def get(self, request):
        # Obtener artículos organizados por categoría
        categorias = Categoria.objects.prefetch_related(
            models.Prefetch(
                'articulo_set',
                queryset=Articulo.objects.filter(activo=True)
                          .select_related('proveedor_principal')
                          .order_by('nombre'),
                to_attr='articulos'
            )
        ).order_by('nombre')

        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Talonario de Pagos"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
        categoria_font = Font(bold=True, color="FFFFFF")
        categoria_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')

        # Encabezado
        ws.merge_cells('A1:H1')
        ws['A1'] = "TALONARIO DE PAGOS - LUBRICANTES LUIS MORA"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_alignment

        ws.merge_cells('A2:H2')
        ws['A2'] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws['A2'].alignment = center_alignment

        # Encabezados de columnas
        columns = [
            ("Código", 15),
            ("Nombre", 40),
            ("Proveedor", 25),
            ("Tipo", 20),
            ("Precio", 15),
            ("IVA %", 10),
            ("Precio con IVA", 15),
            ("Cantidad", 15)
        ]

        # Escribir datos por categoría
        row_num = 4
        for categoria in categorias:
            if not categoria.articulos:
                continue  # Saltar categorías vacías

            # Encabezado de categoría
            ws.merge_cells(f'A{row_num}:H{row_num}')
            ws[f'A{row_num}'] = categoria.nombre.upper()
            ws[f'A{row_num}'].font = categoria_font
            ws[f'A{row_num}'].fill = categoria_fill
            ws[f'A{row_num}'].alignment = center_alignment
            row_num += 1

            # Encabezados de columnas
            for col_num, (column_title, column_width) in enumerate(columns, 1):
                cell = ws.cell(row=row_num, column=col_num, value=column_title)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_alignment
                ws.column_dimensions[get_column_letter(col_num)].width = column_width

            row_num += 1

            # Artículos de la categoría
            for articulo in categoria.articulos:
                precio_con_iva = articulo.precio_venta * (1 + (articulo.tasa_impuesto or 16) / 100)

                ws.cell(row=row_num, column=1, value=articulo.codigo or "SIN-COD").border = thin_border
                ws.cell(row=row_num, column=2, value=articulo.nombre).border = thin_border
                ws.cell(row=row_num, column=3, value=articulo.proveedor_principal.nombre if articulo.proveedor_principal else "N/A").border = thin_border
                ws.cell(row=row_num, column=4, value=articulo.marca or "").border = thin_border
                ws.cell(row=row_num, column=5, value=articulo.precio_venta).border = thin_border
                ws.cell(row=row_num, column=5).number_format = '"COP "#,##0.00'
                ws.cell(row=row_num, column=6, value=articulo.tasa_impuesto*100 or 16).border = thin_border
                ws.cell(row=row_num, column=6).alignment = center_alignment
                ws.cell(row=row_num, column=7, value=precio_con_iva).border = thin_border
                ws.cell(row=row_num, column=7).number_format = '"COP "#,##0.00'
                ws.cell(row=row_num, column=8).border = thin_border
                ws.cell(row=row_num, column=8).alignment = center_alignment

                row_num += 1

            # Espacio entre categorías
            row_num += 1

        # Configurar respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="talonario_pagos_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)

        return response
    
